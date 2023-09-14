import asyncio
import discord
import datetime
import threading
import openpyxl
import time
import random
from discord import Member
from discord.ext import commands
from discord.ext.commands import Bot
from boto3 import client
import urllib
from urllib.request import urlopen, Request
import urllib.request
from selenium import webdriver
import bs4
from discord import opus
from pprint import pprint

if not discord.opus.is_loaded():
	discord.opus.load_opus('opus')
    
    
DiscordClient = discord.Client()

token = "NTc1NTM1NTUwNTI2OTE0NTkw.XNJmHg.zJQSzgt6L_qVNNf3eoTIGVaz2CY" #토큰 입력


#channel = DiscordClient.get_channel('570463651170091029')
channel = DiscordClient.get_channel(580569372511174699)
voice_client = '';
vc = '';
server = ''
chkvoicechannel = 0
yoil = ['월','화','수','목','금','토','일']



async def PlaySound(filename):
    global vc    
    print('play')    
    print(vc)
  
    vc.play(discord.FFmpegPCMAudio(filename), after=lambda e: print('done', e))
    vc.resume()
    while vc.is_playing():
        await asyncio.sleep(1)
    # disconnect after the player has finished
    vc.stop()

    


async def speak(text, filename, lang='ko'):   
    #global server
    print('speak')
    polly = client("polly", aws_access_key_id= "AKIAUU5VCOPTVJFOIN35", aws_secret_access_key="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",region_name="ap-northeast-2")
    response = polly.synthesize_speech(Text=text,OutputFormat="mp3",VoiceId="Seoyeon")
    stream = response.get("AudioStream")
    full_filename = './'+filename+'.mp3'
    with open(full_filename, 'wb') as f:
       data = stream.read()
       f.write(data)
    await asyncio.sleep(1)    
    await PlaySound(full_filename)
    
    
    
    
    
async def my_background_task():
    await DiscordClient.wait_until_ready()
    
    global channel
    channel = DiscordClient.get_channel(580569372511174699)
    
    if channel is not None: 
        while not DiscordClient.is_closed():
        
            print("check")
           
            now = datetime.datetime.now() #지금시간
            
            #priv = datetime.combine(date.today(), time()) + timedelta(hours=1)
            priv = now + datetime.timedelta(minutes=5) #1분전
            priv = priv.time()
            now = now.time()
            today_yoil = yoil[datetime.datetime.today().weekday()]
            
            
            #for a in DiscordClient.get_all_channels()  :
             #   print(a.id)
            #print(now)
            
            file = openpyxl.load_workbook('schedule.xlsx')
            sheet = file.active    
            flag = []
            for i in range(1,201) :
                flag.append(False);
            
            for i in range(1,201) :
                sche_name = sheet["A" + str(i)].value
                sche_time = sheet["B" + str(i)].value
                sche_yoil = sheet["C" + str(i)].value
              
                if sche_name is None :
                    break   
                
            
                
                sche_time = str(sche_time) 
                now = str(now)
                priv = str(priv)
                
                      
                #print (str(sche_name) + str(sche_time) + sche_yoil)
                #print  (sche_time[0:5] + ' ' + priv[0:5] + today_yoil)
                
                if (sche_yoil == '매일' and sche_yoil != '토' and sche_yoil !='일') or sche_yoil == today_yoil:
                    if sche_time[0:5] == now[0:5]:
                      flag[i] = False
                      outstring = '[' + sche_name + '] 시간 입니다.'
                      print(outstring)
                      await speak(outstring,'schedule'+str(i))
                      await channel.send(outstring)

                    
                    if sche_time[0:5] == priv[0:5]:
                        if flag[i] == False:
                            flag[i] = True
                            outstring = '[' + sche_name + '] 시작 5분 전 입니다.'
                            print(outstring)
                            await speak(outstring,'schedule'+str(i))                            
                            await channel.send(outstring )


            await asyncio.sleep(5)

        
        




@DiscordClient.event
async def on_ready():
    global channel
    print("login")
    print(DiscordClient.user.name)
    print(DiscordClient.user.id)
    print("------------------")    
    channel = DiscordClient.get_channel(580569372511174699)
    await DiscordClient.change_presence(status=discord.Status.idle,activity=discord.Activity(name='오늘도 일과 게임'))




@DiscordClient.event
async def on_message(message):
    #print(message)
    if message.author.bot: #만약 메시지를 보낸사람이 봇일 경우에는
        return None #동작하지 않고 무시합니다.
        
    global channel 
    global server
    global voice_client
    global chkvoicechannel
    global member
    global vc
    channel =  message.channel 
    server = message.guild
    
    if message.content.startswith('!현재시간'):
        await channel.send('현재시간은 '+datetime.datetime.now().strftime('%H:%M:%S') + '입니다')
    
    if message.content.startswith('!스케쥴보기'):
        file = openpyxl.load_workbook('schedule.xlsx')
        sheet = file.active    
        text = ''

        for i in range(1, 201):
            if sheet["A" + str(i)].value is None and i == 1:
                text= "데이터없음"
                #await channel.send(channel, "데이터 없음")
            if sheet["A" + str(i)].value is None:
                break
            text = text +"\n" + str(i) + " : ["+str(sheet["A" + str(i)].value) + "]\t" + str(sheet["B" + str(i)].value) + "\t" + str(sheet["C" + str(i)].value)            
            #await channel.send(channel, str(i) + " : ["+str(sheet["A" + str(i)].value) + "] " + str(sheet["C" + str(i)].value) + " " + str(sheet["B" + str(i)].value))
            
        embed = discord.Embed(title="등록된 알림 스케쥴리스트", description=text, color=0x00ff00) 
        embed.set_footer(text =  datetime.datetime.now() )        
        await channel.send(embed=embed)
    
    
    
    
    if message.content.startswith('!말해'):        
        await speak('테스트 중 입니다','test')
        
    if message.content.startswith('!나가'):
        if voice_client is not None:            
            await channel.send('MEGA_BOT이 나갑니다') # 나가드림
            print(vc)
            print(voice_client)
            await voice_client.disconnect()
            chkvoicechannel = 0
    
    
    if message.content.startswith('!소환'):
        voice_client = None
        user=message.author
        
        if user.voice.channel: # can be None
            voice_client=user.voice.channel
            
        print(voice_client)
            
        if voice_client is not None:
            if chkvoicechannel == 0:
                vc = await voice_client.connect()
                print('1')
                print(vc)
                chkvoicechannel = 1                
            else :
                vc = await voice_client.disconnect()
                print('2')
                print(vc)
                voice_client=None
                chkvoicechannel = 0       
                
        else:
            await channel.send('음성채널에 먼저 들어가주세요.')
    

    
    if message.content.startswith('!스케쥴삭제'):
        file = openpyxl.load_workbook('schedule.xlsx')
        sheet = file.active  
        command = message.content.split(" ")        
        title = command[1]
        chk = 0
        
        for i in range(1, 201):
            if sheet["A"+str(i)].value == title: 
                chk = i
                for j in range(i,201) :
                    if sheet["A"+str(j)].value is None:   
                        break;
                    else :
                        sheet["A"+str(j)].value = sheet["A"+str(j+1)].value
                        sheet["B"+str(j)].value = sheet["B"+str(j+1)].value
                        sheet["C"+str(j)].value = sheet["C"+str(j+1)].value
                break
        
        if chk != 0 :        
            file.save("schedule.xlsx")
            await channel.send(str(chk) + ": [" + str(title) + "] 삭제 완료")
        else :
            await channel.send(str(title) + "은 미등록된 스케쥴입니다")
            
    if message.content.startswith('!스케쥴등록'):
        #A 스케쥴제목 
        #B 시간
        #C 요일        
        file = openpyxl.load_workbook('schedule.xlsx')
        sheet = file.active
        command = message.content.split(" ")
        
        #200개 까지 등록 가능
        for i in range(1, 201):
            if sheet["A"+str(i)].value == command[1]: 
                sheet["B" + str(i)].value = command[2]
                
                if len(command) < 4 :                
                    command.append("매일")
                    
                tmpcmd = command[2].split(":")
                
                if len(tmpcmd) < 3 :
                    command[2] = command[2] + ":00"

                sheet["C" + str(i)].value = command[3]
                await channel.send("스케쥴 수정 완료 ["+command[1]+"] "+command[2]+" "+command[3]);
                break
                
            elif sheet["A"+str(i)].value is None:     
                sheet["A"+str(i)].value = command[1]
                sheet["B" + str(i)].value = command[2]
                
                
                if len(command) < 4 :                
                    command.append("매일")
                    
                tmpcmd = command[2].split(":")
                
                if len(tmpcmd) < 3 :
                    command[2] = command[2] + ":00"
 
                    
                sheet["C" + str(i)].value = command[3]
                
                await channel.send("스케쥴 등록 완료 ["+command[1]+"] "+command[2]+" "+command[3]);
                await channel.send("★ 현재 사용중인 데이터 저장용량 : " + str(i)+"/200 ★")
                break
             
                
        file.save("schedule.xlsx")
        
        
    

    if message.content.startswith("!날씨2"):
        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location+'날씨')
        hdr = {'User-Agent': 'Mozilla/5.0'}
        url = 'https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=' + enc_location
        print(url)
        req = Request(url, headers=hdr)
        html = urllib.request.urlopen(req)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        todayBase = bsObj.find('div', {'class': 'main_info'})

        todayTemp1 = todayBase.find('span', {'class': 'todaytemp'})
        todayTemp = todayTemp1.text.strip()  # 온도
        print(todayTemp)

        todayValueBase = todayBase.find('ul', {'class': 'info_list'})
        todayValue2 = todayValueBase.find('p', {'class': 'cast_txt'})
        todayValue = todayValue2.text.strip()  # 밝음,어제보다 ?도 높거나 낮음을 나타내줌
        print(todayValue)

        todayFeelingTemp1 = todayValueBase.find('span', {'class': 'sensible'})
        todayFeelingTemp = todayFeelingTemp1.text.strip()  # 체감온도
        print(todayFeelingTemp)

        todayMiseaMongi1 = bsObj.find('div', {'class': 'sub_info'})
        todayMiseaMongi2 = todayMiseaMongi1.find('div', {'class': 'detail_box'})
        todayMiseaMongi3 = todayMiseaMongi2.find('dd')
        todayMiseaMongi = todayMiseaMongi3.text  # 미세먼지
        print(todayMiseaMongi)

        tomorrowBase = bsObj.find('div', {'class': 'table_info weekly _weeklyWeather'})
        tomorrowTemp1 = tomorrowBase.find('li', {'class': 'date_info'})
        tomorrowTemp2 = tomorrowTemp1.find('dl')
        tomorrowTemp3 = tomorrowTemp2.find('dd')
        tomorrowTemp = tomorrowTemp3.text.strip()  # 오늘 오전,오후온도
        print(tomorrowTemp)

        tomorrowAreaBase = bsObj.find('div', {'class': 'tomorrow_area'})
        tomorrowMoring1 = tomorrowAreaBase.find('div', {'class': 'main_info morning_box'})
        tomorrowMoring2 = tomorrowMoring1.find('span', {'class': 'todaytemp'})
        tomorrowMoring = tomorrowMoring2.text.strip()  # 내일 오전 온도
        print(tomorrowMoring)

        tomorrowValue1 = tomorrowMoring1.find('div', {'class': 'info_data'})
        tomorrowValue = tomorrowValue1.text.strip()  # 내일 오전 날씨상태, 미세먼지 상태
        print(tomorrowValue)

        tomorrowAreaBase = bsObj.find('div', {'class': 'tomorrow_area'})
        tomorrowAllFind = tomorrowAreaBase.find_all('div', {'class': 'main_info morning_box'})
        tomorrowAfter1 = tomorrowAllFind[1]
        tomorrowAfter2 = tomorrowAfter1.find('p', {'class': 'info_temperature'})
        tomorrowAfter3 = tomorrowAfter2.find('span', {'class': 'todaytemp'})
        tomorrowAfterTemp = tomorrowAfter3.text.strip()  # 내일 오후 온도
        print(tomorrowAfterTemp)

        tomorrowAfterValue1 = tomorrowAfter1.find('div', {'class': 'info_data'})
        tomorrowAfterValue = tomorrowAfterValue1.text.strip()

        print(tomorrowAfterValue)  # 내일 오후 날씨상태,미세먼지

        embed = discord.Embed(
            title=learn[1]+ ' 날씨 정보',
            description=learn[1]+ '날씨 정보입니다.',
            colour=discord.Colour.gold()
        )
        embed.add_field(name='현재온도', value=todayTemp+'˚', inline=False)  # 현재온도
        embed.add_field(name='체감온도', value=todayFeelingTemp, inline=False)  # 체감온도
        embed.add_field(name='현재상태', value=todayValue, inline=False)  # 밝음,어제보다 ?도 높거나 낮음을 나타내줌
        embed.add_field(name='현재 미세먼지 상태', value=todayMiseaMongi, inline=False)  # 오늘 미세먼지
        embed.add_field(name='오늘 오전/오후 날씨', value=tomorrowTemp, inline=False)  # 오늘날씨 # color=discord.Color.blue()
        embed.add_field(name='**----------------------------------**',value='**----------------------------------**', inline=False)  # 구분선
        embed.add_field(name='내일 오전온도', value=tomorrowMoring+'˚', inline=False)  # 내일오전날씨
        embed.add_field(name='내일 오전날씨상태, 미세먼지 상태', value=tomorrowValue, inline=False)  # 내일오전 날씨상태
        embed.add_field(name='내일 오후온도', value=tomorrowAfterTemp + '˚', inline=False)  # 내일오후날씨
        embed.add_field(name='내일 오후날씨상태, 미세먼지 상태', value=tomorrowAfterValue, inline=False)  # 내일오후 날씨상태



        await channel.send(embed=embed)
        
DiscordClient.loop.create_task(my_background_task())
DiscordClient.run(token)


